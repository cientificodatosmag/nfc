"""
Extrae de Oracle el listado de modulos de riego con su personal y genera un
Excel crudo, sin limpiar. Para el dato listo para usar, correr despues
limpiar_empleados_modulos.py sobre el archivo que este genera.

Requiere las dependencias: oracledb, openpyxl, keyring.
    pip install oracledb openpyxl keyring

Los datos del servidor salen de tools/conexion.local.json (ver el .ejemplo.json
al lado) y la contrasena del Administrador de Credenciales de Windows. Ninguno
de los dos esta en este archivo: el repositorio es publico.

El Excel que genera lleva nombres de colaboradores, asi que *.xlsx esta en
.gitignore. No lo subas al repositorio.

Que modulos entran
------------------
La tabla NO tiene columna ACTUALIZACION: esa marca solo existia en el Excel
que se compartia a mano. La regla acordada con Ingenio Magdalena es "que tenga
los datos llenos, en especial el numero de ramales", y se comprobo que capturar
NO_RAMALES > 0 recupera los 59 modulos que traia aquel Excel sin perder ninguno.

Los tipos de cantidad fija de etiquetas -hoy la aspersion- entran igual sin
ramales. La regla pedia ramales porque de ellos salia cuantos rotulos grabar; en
aspersion son 6 fijas, asi que exigirlos solo dejaba fuera modulos que se pueden
rotular hoy mismo.

La consulta va dirigida por el MODULO, no por el empleado: asi aparecen tambien
los modulos que cumplen la regla pero todavia no tienen a nadie asignado, que
es justo lo que hay que ver para repartir telefonos. Los colaboradores que no
cuelgan de ningun modulo vigente no se pierden: van en su propia hoja.

CANTIDAD_PERSONAL
-----------------
Columna nueva del maestro: la plantilla que el modulo DEBE tener, incluidas las
plazas todavia sin contratar. No es la plantilla real y no sustituye al join:
solo esta llena en 124 de los 226 modulos (Occidente Centro entera la tiene
vacia), y donde hay gente coincide exactamente con los asignados. Por eso el
reporte la combina con un maximo en vez de creerle a ciegas.

MODULO_ACTIVO_CANICULA
----------------------
Otra columna nueva: SI / No / vacio, 32 / 111 / 84 modulos. Viaja al reporte
como dato para filtrar, no como prioridad. No mide lo mismo que el informe de
equipos: 17 modulos que regaron en los ultimos 7 dias estan marcados 'No', asi
que cambiar una cosa por la otra degradaria modulos que estan operando.
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import _config
from reglas_rotulado import REGLAS

# Regla de "modulo actualizado". Se deja aparte porque es el parametro que hay
# que tocar si algun dia se decide exigir mas campos.
REGLA_ACTUALIZADO = "M.NO_RAMALES IS NOT NULL AND M.NO_RAMALES > 0"

# Los tipos de cantidad fija de etiquetas entran sin ramales: en ellos ese dato
# no dice cuantos rotulos se graban, asi que exigirlo solo deja fuera modulos
# que se pueden rotular. Misma regla y mismo motivo que en actualizar-maestro.py.
TIPOS_SIN_RAMALES = sorted(t for t, (_, _, fijas) in REGLAS.items() if fijas is not None)
REGLA_CANTIDAD_FIJA = (
    "REGEXP_LIKE(M.CODIGO_MODULO, '^[A-Z]{3}-(" + '|'.join(TIPOS_SIN_RAMALES) + ")-[0-9]+$')"
)
REGLA_ENTRADA = f"(({REGLA_ACTUALIZADO}) OR ({REGLA_CANTIDAD_FIJA}))"

COLUMNAS_MODULO = """
    M.ADMIN,
    M.EMPRESA,
    M.C_FINCA,
    M.FINCA,
    M.REGION,
    M.FABRICA,
    M.RESPONSABLE,
    M.TIPO_RIEGO,
    M.FUENTE_AGUA,
    M.ID_MOTOR,
    M.ID_MOTOR_2,
    M.COD_FUNCION,
    M.COD_MAQUINARIA,
    M.CODIGO_MODULO,
    M.NO_HIDRANTES,
    M.NO_RAMALES,
    M.PRESION_OPTIMA,
    M.RPM_OPTIMA,
    M.LAMINA_OPTIMA,
    M.HORAS_RIEGO_PRG,
    M.NO_POSICIONES,
    M.CAUDAL_REQUERIDO,
    M.AREA_MODULO,
    M.CANTIDAD_PERSONAL,
    M.MODULO_ACTIVO_CANICULA,
    M.CREATED_USER,
    M.CREATED_DATE,
    M.LAST_EDITED_USER,
    M.LAST_EDITED_DATE
"""

QUERY = f"""
SELECT
    D.COD_EMPLEADO AS CODIGO_COLABORADOR,
    D.NOM_EMPLEADO AS NOMBRE_COLABORADOR,
{COLUMNAS_MODULO}
FROM
    SDEUSR.MAESTRO_MODULOS_RIEGO M
LEFT OUTER JOIN
    SDEUSR.GRH_EMPLEADOS_MODULOS_RIEGOS D
ON
    M.GLOBALID = D.ID_COD_EMP
WHERE
    {REGLA_ENTRADA}
ORDER BY
    M.REGION, M.RESPONSABLE, M.CODIGO_MODULO, D.NOM_EMPLEADO
"""

# Colaboradores que no cuelgan de ningun modulo que cumpla la regla: o su
# modulo aun no tiene los datos completos, o el GLOBALID no casa con ninguno.
# Se listan aparte para que el cambio de consulta no los haga desaparecer.
QUERY_SIN_MODULO_VIGENTE = f"""
SELECT
    D.COD_EMPLEADO AS CODIGO_COLABORADOR,
    D.NOM_EMPLEADO AS NOMBRE_COLABORADOR,
    M2.CODIGO_MODULO AS MODULO_NO_VIGENTE,
    M2.FINCA,
    M2.REGION,
    M2.NO_RAMALES,
    CASE WHEN M2.CODIGO_MODULO IS NULL
         THEN 'Sin modulo en el maestro'
         ELSE 'El modulo no cumple la regla (ramales vacios y no es de cantidad fija)'
    END AS MOTIVO
FROM
    SDEUSR.GRH_EMPLEADOS_MODULOS_RIEGOS D
LEFT OUTER JOIN
    SDEUSR.MAESTRO_MODULOS_RIEGO M2
ON
    M2.GLOBALID = D.ID_COD_EMP
WHERE
    D.ID_COD_EMP NOT IN (
        SELECT M.GLOBALID FROM SDEUSR.MAESTRO_MODULOS_RIEGO M
        WHERE {REGLA_ENTRADA} AND M.GLOBALID IS NOT NULL
    )
ORDER BY
    D.NOM_EMPLEADO
"""


def escribir_hoja(ws, cols, rows, anchos=None):
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
    for i, col in enumerate(cols, start=1):
        ancho = anchos[i - 1] if anchos else min(max(len(col), 12) + 2, 40)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else "Empleados_Modulos_Riego.xlsx"

    with _config.abrir_oracle() as con:
        cur = con.cursor()

        cur.execute(QUERY)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()

        cur.execute(QUERY_SIN_MODULO_VIGENTE)
        cols_sueltos = [d[0] for d in cur.description]
        sueltos = cur.fetchall()

    i_mod = cols.index("CODIGO_MODULO")
    i_cod = cols.index("CODIGO_COLABORADOR")
    modulos = {r[i_mod] for r in rows}
    con_personal = {r[i_mod] for r in rows if r[i_cod] is not None}

    print(f"Filas obtenidas: {len(rows)}")
    print(f"Modulos que cumplen la regla: {len(modulos)}")
    print(f"  con personal asignado: {len(con_personal)}")
    print(f"  SIN personal asignado: {len(modulos - con_personal)}")
    print(f"Colaboradores sin modulo vigente: {len(sueltos)}")

    wb = Workbook()
    escribir_hoja(wb.active, cols, rows)
    wb.active.title = "Detalle"
    escribir_hoja(wb.create_sheet("Sin modulo vigente"), cols_sueltos, sueltos,
                  anchos=[20, 34, 22, 26, 22, 14, 44])

    wb.save(out_path)
    print(f"Excel crudo generado: {out_path}")
    print("Ahora correr limpiar_empleados_modulos.py sobre este archivo.")


if __name__ == "__main__":
    main()
