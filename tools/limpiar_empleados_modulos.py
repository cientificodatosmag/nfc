"""
Limpia el Excel crudo de extraer_empleados_modulos.py:

- Colapsa espacios sobrantes en NOMBRE_COLABORADOR y RESPONSABLE.
- Quita el prefijo de lista pegado al responsable ("2: Enzo ..." -> "Enzo ...").
- Un mismo CODIGO_COLABORADOR con el nombre escrito distinto entre filas
  (ej. una vez completo y otra solo nombre + primer apellido, o con/sin
  tilde) se unifica a una sola version. Los casos ya revisados estan en
  CORRECCIONES_NOMBRE; si aparece un caso nuevo que este script no conoce,
  se avisa en vez de adivinar.
- Los empleados sin modulo asignado (join sin match en MAESTRO_MODULOS_RIEGO)
  quedan agrupados aparte como "(Sin modulo asignado)" en los resumenes, en
  vez de una fila en blanco.
- Registros con el nombre incompleto en el origen (ORACLE) se marcan en rojo
  con una nota, pero se cuentan igual: revisado con Ingenio Magdalena
  2026-08-03, sí cuentan para el numero de telefonos porque ya tienen
  modulo asignado.

Genera un Excel con: Detalle (limpio), Resumen por modulo, Resumen por
responsable, En varios modulos, y Cambios aplicados (auditoria de que se
toco y por que).

OJO al leer los resumenes: la suma de "Resumen por modulo" es mayor que la
de "Resumen por responsable" porque hay colaboradores que atienden dos
modulos y ahi cuentan una vez por cada uno. La primera suma ASIGNACIONES,
la segunda PERSONAS. Para pedir telefonos vale la segunda. Ambas hojas
llevan su total al pie y la hoja "En varios modulos" lista quienes son.
"""
import os
import re
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import _config

SIN_MODULO = "(Sin modulo asignado)"
SIN_PERSONAL = "(Sin personal asignado)"

# Las decisiones de limpieza viven en tools/normalizaciones.local.json, no aqui:
# son nombres de personas y el repositorio es publico.
#
# - correcciones_nombre: mismo CODIGO_COLABORADOR que llego escrito de forma
#   distinta entre filas. Revisados uno a uno contra el resto de la fila (finca,
#   responsable) para confirmar que es la misma persona antes de unificar.
# - incompletos: nombre que el origen no tiene completo. Cuenta igual para el
#   numero de telefonos (ya tiene modulo), pero se marca para corregirlo.
_NORM = _config.normalizaciones()
CORRECCIONES_NOMBRE = _NORM["correcciones_nombre"]
INCOMPLETOS = _NORM["incompletos"]

ROJO = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
AMARILLO = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
VERDE = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
GRIS = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")

# Etiquetas por modulo: un rotulo por ramal mas cuatro, y dos pasadas completas.
ETIQUETAS_EXTRA = 4
PASADAS = 2

# Niveles de prioridad para rotular, del cruce con el listado de equipos.
#
# Ojo con el matiz: el archivo de canicula es una foto de un instante. Filtrar
# por "Operando" a secas deja fuera a los que estan parados por lluvia, que
# siguen estando en servicio esta temporada. Por eso el nivel 2 existe: la
# senal util no es "esta corriendo ahora" sino "aparece en el listado".
PRIORIDADES = {
    1: "1 - Operando ahora",
    2: "2 - Activo, parado por lluvia",
    3: "3 - Parqueado, otra causa",
    4: "4 - Sin dato en el listado de equipos",
}
RELLENO_PRIORIDAD = {1: VERDE, 2: AMARILLO, 3: AMARILLO, 4: GRIS}


def norm_motor(valor):
    """
    '043-0209', '0043-0209' y '43-209' son el mismo motor.

    El maestro guarda el id con ceros a la izquierda y el listado de equipos sin
    ellos, asi que se comparan como pareja de enteros.
    """
    if valor is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", str(valor).strip())
    return (int(m.group(1)), int(m.group(2))) if m else None


def leer_equipos(ruta):
    """Indexa el listado de canicula por motor. Devuelve {} si no esta el archivo."""
    if not ruta or not os.path.exists(ruta):
        return {}

    ws = openpyxl.load_workbook(ruta, data_only=True).worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    cab = {str(c).strip(): i for i, c in enumerate(filas[0]) if c}

    faltan = [c for c in ("ID", "Estado", "Detalle de Estado") if c not in cab]
    if faltan:
        print(f"AVISO: {ruta} no tiene las columnas {faltan}; se omite el cruce.",
              file=sys.stderr)
        return {}

    equipos = {}
    for fila in filas[1:]:
        clave = norm_motor(fila[cab["ID"]])
        if clave is None:
            continue
        equipos[clave] = {
            "estado": str(fila[cab["Estado"]] or "").strip(),
            "detalle": str(fila[cab["Detalle de Estado"]] or "").strip(),
        }
    return equipos


def clasificar(equipo):
    """Nivel de prioridad a partir del estado del motor que alimenta el modulo."""
    if equipo is None:
        return 4
    if equipo["estado"].lower() == "operando":
        return 1
    if equipo["detalle"].lower() == "lluvia":
        return 2
    return 3


def limpiar_espacios(texto):
    if texto is None:
        return texto
    return re.sub(r"\s+", " ", str(texto).strip())


def limpiar_responsable(texto):
    texto = limpiar_espacios(texto)
    if texto is None:
        return texto
    return re.sub(r"^\d+\s*[:.\-]\s*", "", texto)


def main():
    in_path = sys.argv[1] if len(sys.argv) > 1 else "Empleados_Modulos_Riego.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else in_path.replace(".xlsx", "_LIMPIO.xlsx")
    equipos_path = sys.argv[3] if len(sys.argv) > 3 else "Equipos operando canicula (1).xlsx"

    equipos = leer_equipos(equipos_path)
    if equipos:
        print(f"Listado de equipos cruzado: {len(equipos)} motores desde {equipos_path}")
    else:
        print("Sin cruce de equipos: todos los modulos quedaran en prioridad 4.")

    wb_in = openpyxl.load_workbook(in_path)
    ws_in = wb_in["Detalle"]
    filas = list(ws_in.iter_rows(values_only=True))
    cols = list(filas[0])
    idx = {c: i for i, c in enumerate(cols)}
    datos = [list(f) for f in filas[1:]]

    i_codigo = idx["CODIGO_COLABORADOR"]
    i_nombre = idx["NOMBRE_COLABORADOR"]
    i_resp = idx["RESPONSABLE"]
    i_finca = idx["FINCA"]
    i_modulo = idx["CODIGO_MODULO"]
    i_region = idx["REGION"]
    i_motor = idx["ID_MOTOR"]
    i_motor2 = idx["ID_MOTOR_2"]
    i_ramales = idx["NO_RAMALES"]

    # Aviso de codigos con nombres distintos que este script no conoce
    # todavia, para no corregir a ciegas si aparecen en una extraccion nueva.
    por_codigo = defaultdict(set)
    for fila in datos:
        nombre = limpiar_espacios(fila[i_nombre])
        if nombre:
            por_codigo[fila[i_codigo]].add(nombre)
    nuevos_conflictos = {
        c: v for c, v in por_codigo.items() if len(v) > 1 and c not in CORRECCIONES_NOMBRE
    }
    if nuevos_conflictos:
        print("AVISO: codigos con nombre inconsistente que no estan en CORRECCIONES_NOMBRE:")
        for c, v in nuevos_conflictos.items():
            print(f"  {c}: {sorted(v)}")
        print("Revisalos y agregalos al diccionario antes de confiar en el resultado.\n")

    cambios = []  # (codigo, campo, original, nuevo)
    incompletas_filas = []

    for fila in datos:
        codigo = fila[i_codigo]

        original_nombre = fila[i_nombre]
        nuevo_nombre = limpiar_espacios(original_nombre)
        if codigo in CORRECCIONES_NOMBRE:
            nuevo_nombre = CORRECCIONES_NOMBRE[codigo]
        if nuevo_nombre != (original_nombre or None) and (original_nombre or nuevo_nombre):
            if (original_nombre or "") != (nuevo_nombre or ""):
                cambios.append((codigo, "NOMBRE_COLABORADOR", original_nombre, nuevo_nombre))
        fila[i_nombre] = nuevo_nombre

        original_resp = fila[i_resp]
        nuevo_resp = limpiar_responsable(original_resp)
        if (original_resp or "") != (nuevo_resp or ""):
            cambios.append((codigo, "RESPONSABLE", original_resp, nuevo_resp))
        fila[i_resp] = nuevo_resp

        for campo in (i_finca, i_modulo):
            original = fila[campo]
            nuevo = limpiar_espacios(original)
            if (original or "") != (nuevo or ""):
                cambios.append((codigo, cols[campo], original, nuevo))
            fila[campo] = nuevo

        if not fila[i_modulo]:
            fila[i_finca] = fila[i_finca] or SIN_MODULO
            fila[i_modulo] = fila[i_modulo] or SIN_MODULO
            fila[i_resp] = fila[i_resp] or SIN_MODULO

        # La consulta va dirigida por el modulo, asi que hay filas de modulos
        # que cumplen la regla pero todavia no tienen a nadie. Se marcan para
        # que se vean en vez de pasar como un hueco en blanco.
        if codigo is None:
            fila[i_nombre] = SIN_PERSONAL

        if codigo in INCOMPLETOS:
            incompletas_filas.append(fila)

    # ---------------------------------------------------------------- Detalle
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    cols_detalle = cols + ["ESTADO_EQUIPO", "PRIORIDAD", "OBSERVACION"]
    ws.append(cols_detalle)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for fila in datos:
        nota = INCOMPLETOS.get(fila[i_codigo], "")
        equipo = equipos.get(norm_motor(fila[i_motor]) or norm_motor(fila[i_motor2]))
        nivel = clasificar(equipo)
        ws.append(fila + [equipo["estado"] if equipo else "(no aparece)",
                          PRIORIDADES[nivel], nota])
        if nota:
            for cell in ws[ws.max_row]:
                cell.fill = ROJO
    for i, col in enumerate(cols_detalle, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(col), 12) + 2, 40)
    ws.freeze_panes = "A2"

    # Colaboradores que atienden mas de un modulo. Son la razon por la que la
    # suma de "Resumen por modulo" (asignaciones) es mayor que la de "Resumen
    # por responsable" (personas): en la primera cuentan una vez por modulo.
    # Para pedir telefonos manda el numero de PERSONAS, no el de asignaciones.
    # Un modulo sin personal llega con CODIGO_COLABORADOR nulo. No puede contar
    # como "1 colaborador": se excluye de todos los conteos y en su lugar se
    # informa como modulo sin asignar, que es lo que hay que resolver.
    modulos_por_codigo = defaultdict(set)
    datos_por_codigo = {}
    for fila in datos:
        if fila[i_codigo] is None:
            continue
        modulos_por_codigo[fila[i_codigo]].add(fila[i_modulo])
        datos_por_codigo[fila[i_codigo]] = fila
    compartidos = {c: m for c, m in modulos_por_codigo.items() if len(m) > 1}

    total_personas = len(modulos_por_codigo)

    # ------------------------------------------------------------ Resumen modulo
    conteo_modulo = defaultdict(set)
    for fila in datos:
        clave = (fila[i_resp], fila[i_finca], fila[i_modulo])
        conteo_modulo.setdefault(clave, set())
        if fila[i_codigo] is not None:
            conteo_modulo[clave].add(fila[i_codigo])

    sin_personal = [k for k, v in conteo_modulo.items() if not v]

    ws2 = wb.create_sheet("Resumen por modulo")
    ws2.append(["RESPONSABLE", "FINCA", "CODIGO_MODULO", "NUMERO_COLABORADORES",
                "DE_ESOS_TAMBIEN_EN_OTRO_MODULO"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for (responsable, finca, modulo), colaboradores in sorted(
        conteo_modulo.items(), key=lambda kv: (str(kv[0][0]), str(kv[0][1]), str(kv[0][2]))
    ):
        repetidos = len([c for c in colaboradores if c in compartidos])
        ws2.append([responsable, finca, modulo, len(colaboradores), repetidos])
        if not colaboradores:
            for cell in ws2[ws2.max_row]:
                cell.fill = AMARILLO
    fila_total = ws2.max_row + 2
    ws2.cell(fila_total, 1, "TOTAL asignaciones (suma de la columna)").font = Font(bold=True)
    ws2.cell(fila_total, 4, sum(len(v) for v in conteo_modulo.values())).font = Font(bold=True)
    ws2.cell(fila_total + 1, 1, "TOTAL personas distintas").font = Font(bold=True)
    ws2.cell(fila_total + 1, 4, total_personas).font = Font(bold=True)
    ws2.cell(fila_total + 2, 1, f"Modulos SIN personal asignado (en amarillo): {len(sin_personal)}"
             ).font = Font(bold=True)
    ws2.cell(fila_total + 3, 1,
             f"La diferencia entre las dos cifras son {len(compartidos)} colaboradores que "
             "atienden 2 modulos: aqui cuentan una vez por modulo, pero son una sola persona "
             "(un solo telefono).")
    for i, w in enumerate([38, 30, 22, 22, 30], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # --------------------------------------------------------- Resumen responsable
    conteo_resp = defaultdict(set)
    for fila in datos:
        conteo_resp.setdefault(fila[i_resp], set())
        if fila[i_codigo] is not None:
            conteo_resp[fila[i_resp]].add(fila[i_codigo])

    ws3 = wb.create_sheet("Resumen por responsable")
    ws3.append(["RESPONSABLE", "NUMERO_COLABORADORES"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for responsable, colaboradores in sorted(conteo_resp.items(), key=lambda kv: str(kv[0])):
        ws3.append([responsable, len(colaboradores)])
    fila_total = ws3.max_row + 2
    ws3.cell(fila_total, 1, "TOTAL personas distintas").font = Font(bold=True)
    ws3.cell(fila_total, 2, total_personas).font = Font(bold=True)
    ws3.cell(fila_total + 1, 1,
             "Este es el numero de telefonos a asignar: nadie aparece bajo dos responsables.")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 22
    ws3.freeze_panes = "A2"

    # ----------------------------------------------------------------- Prioridad
    # Un modulo hereda el estado del motor que lo alimenta. Un mismo motor puede
    # alimentar varios modulos (p.ej. CES-MNA-025 y CES-MNA-027 comparten el
    # 0033-0685), asi que la busqueda es por motor, no por modulo.
    estado_modulo = {}
    for fila in datos:
        modulo = fila[i_modulo]
        if modulo in estado_modulo:
            continue
        clave = norm_motor(fila[i_motor]) or norm_motor(fila[i_motor2])
        equipo = equipos.get(clave)
        estado_modulo[modulo] = {
            "motor": fila[i_motor],
            "equipo": equipo,
            "prioridad": clasificar(equipo),
            "ramales": fila[i_ramales],
            "finca": fila[i_finca],
            "responsable": fila[i_resp],
            "region": fila[i_region],
        }

    ws_pr = wb.create_sheet("Prioridad rotulado", 1)
    ws_pr.append(["PRIORIDAD", "CODIGO_MODULO", "FINCA", "RESPONSABLE", "REGION",
                  "ID_MOTOR", "ESTADO_EQUIPO", "DETALLE_ESTADO", "RAMALES",
                  "ETIQUETAS_POR_PASADA", "ETIQUETAS_TOTAL_2_PASADAS", "COLABORADORES"])
    for cell in ws_pr[1]:
        cell.font = Font(bold=True)

    conteo_prioridad = defaultdict(int)
    orden = sorted(estado_modulo.items(),
                   key=lambda kv: (kv[1]["prioridad"], str(kv[0])))
    for modulo, info in orden:
        equipo = info["equipo"]
        ramales = int(info["ramales"]) if info["ramales"] else 0
        por_pasada = ramales + ETIQUETAS_EXTRA if ramales else 0
        personas = len({f[i_codigo] for f in datos
                        if f[i_modulo] == modulo and f[i_codigo] is not None})
        conteo_prioridad[info["prioridad"]] += 1
        ws_pr.append([
            PRIORIDADES[info["prioridad"]], modulo, info["finca"], info["responsable"],
            info["region"], info["motor"],
            equipo["estado"] if equipo else "(no aparece)",
            equipo["detalle"] if equipo else "(no aparece)",
            ramales or None, por_pasada or None,
            (por_pasada * PASADAS) or None, personas,
        ])
        relleno = RELLENO_PRIORIDAD[info["prioridad"]]
        for cell in ws_pr[ws_pr.max_row]:
            cell.fill = relleno

    fila_total = ws_pr.max_row + 2
    ws_pr.cell(fila_total, 1, "RESUMEN").font = Font(bold=True)
    for i, nivel in enumerate(sorted(PRIORIDADES), start=1):
        ws_pr.cell(fila_total + i, 1, PRIORIDADES[nivel]).font = Font(bold=True)
        ws_pr.cell(fila_total + i, 2, conteo_prioridad[nivel]).font = Font(bold=True)
        ws_pr.cell(fila_total + i, 1).fill = RELLENO_PRIORIDAD[nivel]
    ws_pr.cell(fila_total + len(PRIORIDADES) + 1, 1,
               "El listado de equipos es una foto de un instante: la mayoria esta "
               "parada por lluvia, no fuera de servicio. Para priorizar, los niveles "
               "1 y 2 juntos son los modulos activos esta temporada.")
    for i, w in enumerate([32, 20, 26, 34, 20, 16, 16, 22, 12, 22, 26, 16], start=1):
        ws_pr.column_dimensions[get_column_letter(i)].width = w
    ws_pr.freeze_panes = "A2"

    # ------------------------------------------------------------------ Personal
    # Listado unico: una fila por persona, no por asignacion. El carnet es
    # COD_EMPLEADO; la tabla de Oracle no tiene ningun otro identificador.
    por_persona = {}
    for fila in datos:
        codigo = fila[i_codigo]
        if codigo is None:
            continue
        p = por_persona.setdefault(codigo, {
            "nombre": fila[i_nombre], "responsables": set(),
            "regiones": set(), "fincas": set(), "modulos": set()
        })
        p["nombre"] = fila[i_nombre] or p["nombre"]
        for destino, columna in (("responsables", i_resp), ("regiones", i_region),
                                 ("fincas", i_finca), ("modulos", i_modulo)):
            valor = fila[columna]
            if valor:
                p[destino].add(str(valor))

    def juntar(conjunto):
        return ", ".join(sorted(conjunto))

    ws_p = wb.create_sheet("Personal", 1)
    ws_p.append(["CARNET", "NOMBRE_COLABORADOR", "RESPONSABLE", "REGION", "FINCA",
                 "MODULOS_QUE_ATIENDE", "NUM_MODULOS", "OBSERVACION"])
    for cell in ws_p[1]:
        cell.font = Font(bold=True)
    for codigo, p in sorted(por_persona.items(), key=lambda kv: str(kv[1]["nombre"]).upper()):
        nota = INCOMPLETOS.get(codigo, "")
        ws_p.append([codigo, p["nombre"], juntar(p["responsables"]), juntar(p["regiones"]),
                     juntar(p["fincas"]), juntar(p["modulos"]), len(p["modulos"]), nota])
        if nota:
            for cell in ws_p[ws_p.max_row]:
                cell.fill = ROJO
    fila_total = ws_p.max_row + 2
    ws_p.cell(fila_total, 1, "TOTAL personas distintas").font = Font(bold=True)
    ws_p.cell(fila_total, 7, len(por_persona)).font = Font(bold=True)
    ws_p.cell(fila_total + 1, 1,
              "Una fila por persona: este es el numero de telefonos a asignar. "
              "El carnet es el codigo de empleado de Oracle (COD_EMPLEADO).")
    for i, w in enumerate([14, 34, 34, 20, 30, 30, 14, 44], start=1):
        ws_p.column_dimensions[get_column_letter(i)].width = w
    ws_p.freeze_panes = "A2"

    # ------------------------------------------------- Colaboradores compartidos
    ws5 = wb.create_sheet("En varios modulos")
    ws5.append(["CODIGO_COLABORADOR", "NOMBRE_COLABORADOR", "RESPONSABLE",
                "MODULOS_QUE_ATIENDE", "CUANTOS_MODULOS"])
    for cell in ws5[1]:
        cell.font = Font(bold=True)
    for codigo, mods in sorted(compartidos.items()):
        fila = datos_por_codigo[codigo]
        ws5.append([codigo, fila[i_nombre], fila[i_resp],
                    ", ".join(sorted(str(m) for m in mods)), len(mods)])
    for i, w in enumerate([20, 34, 34, 30, 18], start=1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    ws5.freeze_panes = "A2"

    # ------------------------------------------------------------------- Auditoria
    ws4 = wb.create_sheet("Cambios aplicados")
    ws4.append(["CODIGO_COLABORADOR", "CAMPO", "VALOR_ORIGINAL", "VALOR_NUEVO"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for codigo, campo, original, nuevo in cambios:
        ws4.append([codigo, campo, original, nuevo])
    for i, w in enumerate([20, 20, 35, 35], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"

    # La hoja de colaboradores sin modulo vigente se arrastra tal cual desde el
    # crudo: no se limpia, pero tampoco se pierde al pasar por aqui.
    if "Sin modulo vigente" in wb_in.sheetnames:
        origen = wb_in["Sin modulo vigente"]
        ws6 = wb.create_sheet("Sin modulo vigente")
        for fila in origen.iter_rows(values_only=True):
            ws6.append(list(fila))
        for cell in ws6[1]:
            cell.font = Font(bold=True)
        for i, w in enumerate([20, 34, 22, 26, 22, 14, 44], start=1):
            ws6.column_dimensions[get_column_letter(i)].width = w
        ws6.freeze_panes = "A2"

    try:
        wb.save(out_path)
    except PermissionError:
        print(f"No se pudo escribir {out_path}: Windows lo tiene bloqueado.", file=sys.stderr)
        print("Seguramente esta abierto en Excel. Cierralo y vuelve a correr el script.",
              file=sys.stderr)
        sys.exit(1)

    modulos = {fila[i_modulo] for fila in datos}
    print(f"Filas procesadas (asignaciones): {len(datos)}")
    print(f"Modulos que cumplen la regla: {len(modulos)}")
    print(f"  SIN personal asignado: {len(sin_personal)}")
    for nivel in sorted(PRIORIDADES):
        print(f"  {PRIORIDADES[nivel]}: {conteo_prioridad[nivel]}")
    print(f"Personas distintas (telefonos): {total_personas}")
    print(f"Colaboradores en 2+ modulos: {len(compartidos)}")
    print(f"Cambios de texto aplicados: {len(cambios)}")
    print(f"Registros marcados incompletos: {len(incompletas_filas)}")
    print(f"Excel limpio generado: {out_path}")


if __name__ == "__main__":
    main()
