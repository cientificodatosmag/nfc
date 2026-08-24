"""
Propone CODIGO_MODULO para las filas que Oracle tiene con datos pero sin codigo.

    python tools/proponer_codigos_modulo.py
    python tools/proponer_codigos_modulo.py --equipos "INFORME_EQUIPO_RIEGO (3).xlsx"

Estas filas existen en SDEUSR.MAESTRO_MODULOS_RIEGO con finca, motor, ramales e
hidrantes llenos, pero con el CODIGO_MODULO vacio. Sin codigo no hay nada que
grabar en la etiqueta, asi que actualizar-maestro.py las descarta y solo dice
cuantas son. Este script arma la propuesta para que alguien las capture en
Oracle; NO escribe en Oracle ni toca el maestro.

Como se arma el codigo: ZONA-TIPO-CORRELATIVO
---------------------------------------------
La ZONA sale de la REGION, con el mapa deducido de los modulos que ya tienen
codigo (CENTRAL NORTE -> CEN, CENTRAL SUR -> CES, ...) en vez de escrito a mano:
si manana aparece una region nueva, el script avisa en vez de inventarla.

El TIPO sale de TIPO_RIEGO cuando Oracle lo trae. Cuando viene vacio -12 de las
24 filas- se cruza el ID_MOTOR contra INFORME_EQUIPO_RIEGO, que si sabe que
riega ese motor. La equivalencia entre el nombre del informe y la sigla tampoco
esta escrita a mano: se deduce de los modulos que YA tienen codigo y motor, y la
columna ORIGEN_TIPO dice de donde salio cada una. Donde las dos fuentes se
contradicen, el Excel lo dice en vez de elegir en silencio.

El CORRELATIVO es el ultimo de esa serie mas uno, como se pidio. "El ultimo" se
busca en TRES sitios, no solo en Oracle: Oracle, modulos.json y el registro
compartido. La razon es que hay codigos que Oracle ya no tiene pero cuyas
etiquetas estan pegadas en el campo -los que el maestro conserva-, y reusar uno
de esos numeros daria dos modulos distintos con el mismo rotulo.

Los huecos intermedios NO se rellenan: CES-MNA salta del 20 al 22 y el siguiente
que propone este script es el 81, no el 21. Se pidio asi, y ademas un hueco casi
siempre es un codigo retirado, no un numero libre.

Lo que NO se propone
--------------------
Las filas que corresponden a un modulo que ya estuvo en la app y perdio el
codigo van a su propia hoja, sin codigo nuevo. Darles uno crearia un segundo
codigo para etiquetas que ya dicen otra cosa en el campo. Se detectan por finca
+ tipo + ramales contra los modulos que el maestro conserva, y ademas se avisa
de cualquier fila que caiga en una finca donde hay huerfanos, aunque no calce
exacto: ahi la coincidencia hay que mirarla con el maestro delante.
"""
import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import _config
from reglas_rotulado import NOMBRES, REGLAS, por_pasada, pasadas_de

RAIZ = Path(__file__).resolve().parent.parent

CODIGO = re.compile(r"^([A-Z]{3})-([A-Z]{3})-(\d+)$")

AMARILLO = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
ROJO = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
GRIS = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")

QUERY_SIN_CODIGO = """
SELECT REGION, FINCA, C_FINCA, RESPONSABLE, TIPO_RIEGO, FUENTE_AGUA,
       ID_MOTOR, ID_MOTOR_2, NO_RAMALES, NO_HIDRANTES, AREA_MODULO,
       CANTIDAD_PERSONAL, MODULO_ACTIVO_CANICULA, OBJECTID
FROM SDEUSR.MAESTRO_MODULOS_RIEGO
WHERE CODIGO_MODULO IS NULL OR TRIM(CODIGO_MODULO) IS NULL
ORDER BY REGION, FINCA, ID_MOTOR
"""

QUERY_CON_CODIGO = """
SELECT CODIGO_MODULO, REGION, TIPO_RIEGO, ID_MOTOR, ID_MOTOR_2
FROM SDEUSR.MAESTRO_MODULOS_RIEGO
WHERE REGEXP_LIKE(CODIGO_MODULO, '^[A-Z]{3}-[A-Z]{3}-[0-9]+$')
"""


def norm_motor(valor):
    """'043-0209', '0043-0209' y '43-209' son el mismo motor. Igual que en los reportes."""
    if valor is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", str(valor).strip())
    return (int(m.group(1)), int(m.group(2))) if m else None


def texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor == int(valor):
        return str(int(valor))
    return re.sub(r"\s+", " ", str(valor)).strip()


def entero(valor):
    return int(valor) if valor is not None else None


def clave_tipo(valor):
    """'Mini Aspersion', 'Miniaspersion' y 'MINIASPERSION' son lo mismo."""
    t = texto(valor).lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        t = t.replace(a, b)
    return re.sub(r"[^a-z]", "", t)


# --------------------------------------------------------------------- fuentes

def leer_equipos(ruta):
    """
    Por motor: que riega, con que funcion y quien responde, del informe de campo.

    Se queda con la ultima fila de cada motor, que es la foto mas reciente. La
    funcion viaja porque un motor 'Abastecedor' alimenta a otros y no siempre es
    un modulo que haya que rotular: es un dato para que lo mire quien captura.
    """
    if not Path(ruta).exists():
        return {}
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["BD"] if "BD" in wb.sheetnames else wb.worksheets[0]
    filas = ws.iter_rows(values_only=True)
    cab = {texto(c): i for i, c in enumerate(next(filas, ()) or ()) if c}
    col_funcion = next((k for k in cab if k.lower().startswith("funci")), None)
    if "ID" not in cab or "Tipo de Riego" not in cab:
        return {}

    equipos = {}
    for fila in filas:
        k = norm_motor(fila[cab["ID"]])
        if k is None:
            continue
        equipos[k] = {
            "tipo": texto(fila[cab["Tipo de Riego"]]),
            "funcion": texto(fila[cab[col_funcion]]) if col_funcion else "",
            "responsable": texto(fila[cab["Responsable"]]) if "Responsable" in cab else "",
        }
    return equipos


def codigos_ya_usados():
    """
    Todo codigo que exista en cualquier parte, para no reusar un correlativo.

    Oracle solo no basta: los modulos que el maestro conserva perdieron su fila
    en Oracle pero tienen etiquetas pegadas en el campo con ese codigo escrito.
    """
    usados = defaultdict(set)
    fuentes = Counter()

    def anotar(codigo, fuente):
        m = CODIGO.match(texto(codigo).upper())
        if m:
            usados[(m.group(1), m.group(2))].add(int(m.group(3)))
            fuentes[fuente] += 1

    with _config.abrir_oracle() as con:
        cur = con.cursor()
        cur.execute("SELECT DISTINCT CODIGO_MODULO FROM SDEUSR.MAESTRO_MODULOS_RIEGO "
                    "WHERE CODIGO_MODULO IS NOT NULL")
        for (c,) in cur:
            anotar(c, "oracle")

    maestro = RAIZ / "modulos.json"
    if maestro.exists():
        d = json.loads(maestro.read_text(encoding="utf-8"))
        for m in (d["modulos"] if isinstance(d, dict) else d):
            anotar(m.get("codigo"), "maestro")

    resumen = RAIZ / "rotulado_resumen.csv"
    if resumen.exists():
        import csv
        with resumen.open(encoding="utf-8-sig", newline="") as fh:
            for r in csv.DictReader(fh):
                anotar(r.get("Codigo_modulo"), "registro")

    return usados, fuentes


def mapas_desde_los_ya_codificados(con_codigo, equipos):
    """
    REGION -> zona y nombre de tipo -> sigla, deducidos y no escritos a mano.

    Se aprende de los modulos que ya tienen codigo: su propia sigla dice a que
    equivale el texto que traen Oracle y el informe. Asi el dia que aparezca una
    region o un tipo nuevo, el script lo dice en vez de adivinar.
    """
    zona = defaultdict(Counter)
    tipo_oracle = defaultdict(Counter)
    tipo_informe = defaultdict(Counter)

    for codigo, region, tipo, m1, m2 in con_codigo:
        m = CODIGO.match(texto(codigo).upper())
        if not m:
            continue
        pref, sigla = m.group(1), m.group(2)
        if texto(region):
            zona[texto(region).upper()][pref] += 1
        if clave_tipo(tipo):
            tipo_oracle[clave_tipo(tipo)][sigla] += 1
        eq = equipos.get(norm_motor(m1) or norm_motor(m2))
        if eq and clave_tipo(eq["tipo"]):
            tipo_informe[clave_tipo(eq["tipo"])][sigla] += 1

    def elegir(tabla):
        salida = {}
        for k, cuenta in tabla.items():
            sigla, n = cuenta.most_common(1)[0]
            salida[k] = (sigla, n / sum(cuenta.values()), dict(cuenta))
        return salida

    return elegir(zona), elegir(tipo_oracle), elegir(tipo_informe)


def huerfanos_del_maestro():
    """Modulos que el maestro conserva porque Oracle les vacio el codigo."""
    maestro = RAIZ / "modulos.json"
    if not maestro.exists():
        return []
    d = json.loads(maestro.read_text(encoding="utf-8"))
    ms = d["modulos"] if isinstance(d, dict) else d
    return [m for m in ms if m.get("conservado")]


# ---------------------------------------------------------------------- salida

def encabezado(ws, cols, anchos):
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--equipos", default="INFORME_EQUIPO_RIEGO (3).xlsx",
                    help="informe de equipos, para el tipo de riego que Oracle no trae")
    ap.add_argument("--salida", default="")
    args = ap.parse_args()

    equipos = leer_equipos(RAIZ / args.equipos if not Path(args.equipos).is_absolute()
                           else args.equipos)
    print(f"Informe de equipos: {len(equipos)} motores desde {args.equipos}"
          if equipos else f"AVISO: no se pudo leer {args.equipos}; "
                          "las filas sin TIPO_RIEGO se quedaran sin sigla.")

    with _config.abrir_oracle() as con:
        cur = con.cursor()
        cur.execute(QUERY_SIN_CODIGO)
        cols = [d[0] for d in cur.description]
        sin_codigo = [dict(zip(cols, f)) for f in cur]
        cur.execute(QUERY_CON_CODIGO)
        con_codigo = cur.fetchall()

    print(f"Filas en Oracle sin CODIGO_MODULO: {len(sin_codigo)}")

    zonas, t_oracle, t_informe = mapas_desde_los_ya_codificados(con_codigo, equipos)
    usados, fuentes = codigos_ya_usados()
    print(f"Codigos ya usados: {sum(fuentes.values())} "
          f"({', '.join(f'{k} {v}' for k, v in fuentes.items())})")

    huerfanos = huerfanos_del_maestro()
    por_finca = defaultdict(list)
    for h in huerfanos:
        por_finca[clave_tipo(h["finca"])].append(h)

    propuestas, revisar, sin_sigla = [], [], []

    for fila in sin_codigo:
        region = texto(fila["REGION"])
        finca = texto(fila["FINCA"])
        ramales = entero(fila["NO_RAMALES"])
        motor = texto(fila["ID_MOTOR"]) or texto(fila["ID_MOTOR_2"])
        eq = equipos.get(norm_motor(fila["ID_MOTOR"]) or norm_motor(fila["ID_MOTOR_2"]))

        # ---- sigla del tipo de riego
        sigla_or = t_oracle.get(clave_tipo(fila["TIPO_RIEGO"]), (None,))[0]
        sigla_in = t_informe.get(clave_tipo(eq["tipo"]), (None,))[0] if eq else None
        if sigla_or:
            sigla, origen = sigla_or, "Oracle (TIPO_RIEGO)"
        elif sigla_in:
            sigla, origen = sigla_in, "informe de equipos (por motor)"
        else:
            sigla, origen = None, "SIN DATO"

        conflicto = ""
        if sigla_or and sigla_in and sigla_or != sigla_in:
            conflicto = (f"Oracle dice {sigla_or} y el informe {sigla_in} "
                         f"('{texto(fila['TIPO_RIEGO'])}' vs '{eq['tipo']}')")

        # ---- corresponde a un modulo que ya estuvo en la app?
        candidatos = por_finca.get(clave_tipo(finca), [])
        exacto = next((h for h in candidatos
                       if ramales is not None and h.get("ramales") == ramales), None)
        registro = {
            "region": region, "zona": zonas.get(region.upper(), ("?",))[0],
            "finca": finca, "c_finca": texto(fila["C_FINCA"]),
            "responsable": texto(fila["RESPONSABLE"]) or (eq["responsable"] if eq else ""),
            "tipo_oracle": texto(fila["TIPO_RIEGO"]),
            "tipo_informe": eq["tipo"] if eq else "",
            "funcion": eq["funcion"] if eq else "",
            "sigla": sigla, "origen": origen, "conflicto": conflicto,
            "motor": motor, "motor2": texto(fila["ID_MOTOR_2"]),
            "ramales": ramales, "hidrantes": entero(fila["NO_HIDRANTES"]),
            "area": fila["AREA_MODULO"], "personal": entero(fila["CANTIDAD_PERSONAL"]),
            "canicula": texto(fila["MODULO_ACTIVO_CANICULA"]),
            "objectid": entero(fila["OBJECTID"]),
        }

        if exacto:
            registro["huerfano"] = exacto["codigo"]
            registro["nota"] = (
                f"Coincide con {exacto['codigo']} (misma finca, {ramales} ramales), "
                f"que ya estuvo en la app. NO darle codigo nuevo sin revisar: "
                f"sus etiquetas estan pegadas en el campo.")
            revisar.append(registro)
            continue

        if candidatos:
            registro["nota"] = (
                f"La finca tiene {len(candidatos)} modulo(s) huerfano(s) "
                f"({', '.join(h['codigo'] for h in candidatos)}). No calza por ramales, "
                f"pero conviene mirarlo antes de capturar.")
        else:
            registro["nota"] = ""

        if sigla is None:
            sin_sigla.append(registro)
        propuestas.append(registro)

    # ---- correlativos: el ultimo de la serie mas uno, en orden estable
    propuestas.sort(key=lambda r: (r["region"], r["finca"], r["motor"]))
    siguiente = {}
    for r in propuestas:
        if not r["sigla"] or r["zona"] == "?":
            r["codigo"] = ""
            continue
        k = (r["zona"], r["sigla"])
        if k not in siguiente:
            siguiente[k] = max(usados.get(k, {0})) + 1
        r["codigo"] = f"{r['zona']}-{r['sigla']}-{siguiente[k]:03d}"
        r["ultimo_previo"] = max(usados.get(k, {0}))
        siguiente[k] += 1

    # ---- etiquetas que llevaria cada uno, con la regla del repo
    #
    # Se deja en blanco antes que poner un numero que no significa nada, en dos
    # casos que hay en los datos de hoy:
    #
    #  - Tipos que la app no rotula (BGR, GRA, APO, SRI). Caen en la regla por
    #    defecto y saldrian con "ramales + 4, dos juegos" como si fueran mini
    #    aspersion. Un bombeo por gravedad no lleva rotulos de riego.
    #  - Tipos que cuentan por ramal y llegan SIN ramales. La cuenta daria el
    #    puro extra -0 + 4, dos juegos = 8- y ese 8 es un artefacto de la
    #    formula, no una cantidad que alguien deba mandar a imprimir.
    for r in propuestas:
        r["por_juego"] = r["juegos"] = r["total"] = ""
        if not r["codigo"]:
            continue
        if r["sigla"] not in REGLAS:
            r["nota"] = (f"{r['sigla']} no es un tipo que la app rotule; no se calculan "
                         f"etiquetas. " + r["nota"]).strip()
            continue
        cuenta_por_ramal = REGLAS[r["sigla"]][2] is None
        if cuenta_por_ramal and not r["ramales"]:
            r["nota"] = ("Sin NO_RAMALES en Oracle no se puede saber cuantas etiquetas "
                         "lleva: hay que llenar ese dato. " + r["nota"]).strip()
            continue
        r["por_juego"] = por_pasada(r["codigo"], r["ramales"] or 0)
        r["juegos"] = pasadas_de(r["codigo"])
        r["total"] = r["por_juego"] * r["juegos"]

    # ------------------------------------------------------------------ Excel
    destino = Path(args.salida or
                   RAIZ / f"Codigos_Propuestos_{date.today().strftime('%Y%m%d')}.xlsx")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Codigos propuestos"
    COLS = ["CODIGO_PROPUESTO", "ZONA", "REGION", "TIPO_RIEGO_SIGLA", "TIPO_RIEGO",
            "ORIGEN_DEL_TIPO", "FINCA", "C_FINCA", "ID_MOTOR", "ID_MOTOR_2",
            "RESPONSABLE", "FUNCION_MOTOR", "NO_RAMALES", "NO_HIDRANTES",
            "AREA_MODULO", "CANTIDAD_PERSONAL", "ACTIVO_CANICULA",
            "ETIQUETAS_POR_JUEGO", "JUEGOS", "ETIQUETAS_TOTAL",
            "ULTIMO_DE_LA_SERIE", "CONFLICTO_DE_TIPO", "NOTA", "OBJECTID"]
    encabezado(ws, COLS, [19, 7, 17, 9, 20, 26, 22, 9, 12, 12, 26, 14, 8, 8,
                          11, 9, 9, 10, 7, 10, 10, 46, 60, 10])
    for r in propuestas:
        ws.append([
            r["codigo"], r["zona"], r["region"], r["sigla"] or "",
            NOMBRES.get(r["sigla"], r["tipo_oracle"] or r["tipo_informe"]),
            r["origen"], r["finca"], r["c_finca"], r["motor"], r["motor2"],
            r["responsable"], r["funcion"], r["ramales"], r["hidrantes"],
            r["area"], r["personal"], r["canicula"],
            r["por_juego"], r["juegos"], r["total"],
            r.get("ultimo_previo", ""), r["conflicto"], r["nota"], r["objectid"],
        ])
        if not r["codigo"]:
            for c in ws[ws.max_row]:
                c.fill = ROJO
        elif r["conflicto"] or r["nota"]:
            for c in ws[ws.max_row]:
                c.fill = AMARILLO

    ws2 = wb.create_sheet("Revisar - ya estuvieron")
    encabezado(ws2, ["MODULO_QUE_LE_CORRESPONDE", "ZONA", "REGION", "FINCA",
                     "TIPO_RIEGO", "ID_MOTOR", "RESPONSABLE", "NO_RAMALES",
                     "NO_HIDRANTES", "CANTIDAD_PERSONAL", "NOTA", "OBJECTID"],
               [26, 7, 17, 22, 20, 12, 26, 9, 9, 9, 74, 10])
    for r in sorted(revisar, key=lambda x: x["huerfano"]):
        ws2.append([r["huerfano"], r["zona"], r["region"], r["finca"],
                    r["tipo_oracle"] or r["tipo_informe"], r["motor"],
                    r["responsable"], r["ramales"], r["hidrantes"],
                    r["personal"], r["nota"], r["objectid"]])
        for c in ws2[ws2.max_row]:
            c.fill = GRIS

    ws3 = wb.create_sheet("Series y correlativos")
    encabezado(ws3, ["SERIE", "ULTIMO_USADO", "CUANTOS_USADOS", "PRIMER_PROPUESTO",
                     "CUANTOS_SE_PROPONEN", "HUECOS_QUE_NO_SE_RELLENAN"],
               [12, 13, 15, 17, 20, 44])
    por_serie = Counter((r["zona"], r["sigla"]) for r in propuestas if r["codigo"])
    for k in sorted(por_serie):
        n = usados.get(k, {0})
        mx = max(n)
        huecos = sorted(set(range(1, mx + 1)) - n)
        ws3.append([f"{k[0]}-{k[1]}", mx, len(n - {0}), f"{mx + 1:03d}", por_serie[k],
                    ", ".join(f"{h:03d}" for h in huecos) or "(ninguno)"])

    wb.save(destino)

    # ------------------------------------------------------------------ consola
    print()
    print(f"Codigos propuestos : {sum(1 for r in propuestas if r['codigo'])}")
    print(f"Sin sigla (en rojo): {len(sin_sigla)}")
    print(f"Para revisar       : {len(revisar)} (corresponden a un modulo que ya estuvo)")
    conflictos = [r for r in propuestas if r["conflicto"]]
    if conflictos:
        print(f"Conflictos de tipo : {len(conflictos)}")
        for r in conflictos:
            print(f"    {r['codigo'] or '(sin codigo)'}  {r['finca']}: {r['conflicto']}")
    print()
    for k in sorted(por_serie):
        print(f"    {k[0]}-{k[1]}: {por_serie[k]} codigo(s) desde "
              f"{max(usados.get(k, {0})) + 1:03d}")
    print()
    print(f"Excel generado: {destino}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
