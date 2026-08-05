"""
Anade a un Excel de modulos el avance real de rotulado que llevan los telefonos.

    python tools/avance_en_excel.py Modulos_Riego_Prioridad_20260804.xlsx

No consulta Oracle ni regenera nada: parte del Excel que ya existe y le pega
encima lo que dice el registro compartido. Escribe un archivo nuevo con la fecha
de hoy y no toca el original.

Lo que anade:

  - Columnas nuevas en la hoja de prioridad: cuantas etiquetas lleva cada pasada
    y en que estado esta el modulo (Completo / En proceso / Sin empezar).
  - Una hoja "Avance rotulado" con la foto completa, ordenada por lo que falta.

Que cuenta como completo
------------------------
Las DOS pasadas llenas. Un modulo con la pasada 1 al 100%% y la 2 a cero esta a
la mitad, no terminado: son dos juegos de etiquetas fisicas distintos.

Las etiquetas que hacen falta salen de ramales + 4, igual que en la app. Si el
Excel y el maestro no coinciden en los ramales de un modulo, manda el maestro,
que es lo que tienen los telefonos delante.
"""
import json
import os
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
ETIQUETAS_EXTRA = 4
PASADAS = 2

VERDE = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
AMARILLO = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
ROJO = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
GRIS = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")

COMPLETO = "Completo"
EN_PROCESO = "En proceso"
SIN_EMPEZAR = "Sin empezar"

RELLENO_ESTADO = {COMPLETO: VERDE, EN_PROCESO: AMARILLO, SIN_EMPEZAR: GRIS}


# --------------------------------------------------------------- el registro

def base_url():
    cfg = RAIZ / "config.js"
    for linea in cfg.read_text(encoding="utf-8").splitlines():
        if "apiBase" in linea and "'" in linea:
            return linea.split("'")[1].rstrip("/")
    sys.exit("No se encontro apiBase en config.js")


def bajar_eventos():
    llave = os.environ.get("NFC_APP_KEY")
    if not llave:
        sys.exit("Falta NFC_APP_KEY en el entorno: es lo que da acceso al registro.")

    eventos, desde = [], 0
    for _ in range(50):
        pet = urllib.request.Request(
            f"{base_url()}/api/eventos?desde={desde}&limite=1000",
            headers={"x-app-key": llave})
        try:
            with urllib.request.urlopen(pet, timeout=60) as r:
                datos = json.loads(r.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError) as e:
            sys.exit(f"No se pudo leer el registro compartido: {e}")
        eventos.extend(datos.get("eventos", []))
        if not datos.get("hayMas"):
            break
        desde = datos.get("siguienteCursor", desde)
    return eventos


def proyectar(eventos):
    """
    Misma regla que el servidor y que la app: barreras de reinicio, y a igual
    clave gana la fecha mayor. Tener aqui una tercera version de esto seria
    pedir que algun dia dijeran cosas distintas, pero es de lectura y da igual
    quien gane: lo que importa es cuantas claves distintas quedan vivas.
    """
    barreras = defaultdict(dict)
    for e in eventos:
        if e.get("tipo") == "reset":
            previa = barreras[e["modulo"]].get(e["pasada"])
            if not previa or e["fecha"] > previa:
                barreras[e["modulo"]][e["pasada"]] = e["fecha"]

    vivas = {}
    for e in eventos:
        if e.get("tipo") == "reset":
            continue
        corte = barreras.get(e["modulo"], {}).get(e["pasada"])
        if corte and e["fecha"] <= corte:
            continue
        clave = (e["modulo"], e["pasada"], e["numero"])
        previo = vivas.get(clave)
        if not previo or e["fecha"] > previo["fecha"]:
            vivas[clave] = e
    return vivas


def resumir(vivas):
    """Por modulo: cuantas de cada pasada, quien las grabo y cuando."""
    por_modulo = defaultdict(lambda: {
        1: 0, 2: 0, "ultima": "", "telefonos": set(), "duplicadas": 0,
    })
    dispositivos = defaultdict(set)

    for (modulo, pasada, numero), e in vivas.items():
        dispositivos[(modulo, pasada, numero)].add(e["dispositivo"])

    for (modulo, pasada, numero), e in vivas.items():
        m = por_modulo[modulo]
        m[pasada] = m.get(pasada, 0) + 1
        m["telefonos"].add(e["dispositivo"])
        if e["fecha"] > m["ultima"]:
            m["ultima"] = e["fecha"]
    return por_modulo


def duplicados_por_modulo(eventos, vivas):
    """Claves que dos telefonos distintos grabaron: hay dos etiquetas iguales."""
    quien = {}
    repetidas = defaultdict(set)
    for e in eventos:
        if e.get("tipo") == "reset":
            continue
        clave = (e["modulo"], e["pasada"], e["numero"])
        if clave not in vivas:
            continue
        if clave in quien and quien[clave] != e["dispositivo"]:
            repetidas[e["modulo"]].add(clave)
        quien[clave] = e["dispositivo"]
    return {m: len(v) for m, v in repetidas.items()}


# ------------------------------------------------------------------ maestro

def ramales_del_maestro():
    ruta = RAIZ / "modulos.json"
    if not ruta.exists():
        return {}
    datos = json.loads(ruta.read_text(encoding="utf-8"))
    return {m["codigo"]: m for m in datos.get("modulos", [])}


def estado_de(hechas1, hechas2, por_pasada):
    if por_pasada <= 0:
        return SIN_EMPEZAR
    if hechas1 >= por_pasada and hechas2 >= por_pasada:
        return COMPLETO
    if hechas1 + hechas2 == 0:
        return SIN_EMPEZAR
    return EN_PROCESO


# -------------------------------------------------------------------- excel

def ancho(ws):
    for col in range(1, ws.max_column + 1):
        largo = max((len(str(ws.cell(row=f, column=col).value or ""))
                     for f in range(1, min(ws.max_row, 400) + 1)), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(max(largo + 2, 10), 42)


def main():
    origen = Path(sys.argv[1] if len(sys.argv) > 1
                  else "Modulos_Riego_Prioridad_20260804.xlsx")
    if not origen.is_absolute():
        origen = RAIZ / origen
    if not origen.exists():
        sys.exit(f"No existe {origen}")

    destino = RAIZ / f"{origen.stem.rsplit('_', 1)[0]}_{date.today():%Y%m%d}.xlsx"
    if destino == origen:
        destino = origen.with_name(f"{origen.stem}_CON_AVANCE.xlsx")

    print(f"Excel de partida : {origen.name}")
    eventos = bajar_eventos()
    vivas = proyectar(eventos)
    avance = resumir(vivas)
    repetidas = duplicados_por_modulo(eventos, vivas)
    maestro = ramales_del_maestro()
    print(f"Registro         : {len(eventos)} eventos, {len(vivas)} etiquetas vigentes")
    print(f"Modulos tocados  : {len(avance)}")

    wb = openpyxl.load_workbook(origen)

    # ---------------------------------------------- columnas en la hoja de prioridad
    hoja_pri = next((wb[n] for n in wb.sheetnames if "prioridad" in n.lower()), None)
    modulos_excel = {}

    if hoja_pri is not None:
        cab = {str(c.value).strip(): c.column for c in hoja_pri[1] if c.value}
        col_mod = cab.get("CODIGO_MODULO")
        col_ram = cab.get("RAMALES") or cab.get("NO_RAMALES")
        if col_mod:
            inicio = hoja_pri.max_column + 1
            nuevas = ["P1_GRABADAS", "P2_GRABADAS", "FALTAN", "ESTADO_ROTULADO", "ULTIMA_FECHA"]
            for i, nombre in enumerate(nuevas):
                celda = hoja_pri.cell(row=1, column=inicio + i, value=nombre)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")

            for fila in range(2, hoja_pri.max_row + 1):
                codigo = hoja_pri.cell(row=fila, column=col_mod).value
                if not isinstance(codigo, str) or codigo.count("-") != 2:
                    continue
                ram = maestro.get(codigo, {}).get("ramales")
                if ram is None and col_ram:
                    valor = hoja_pri.cell(row=fila, column=col_ram).value
                    ram = int(valor) if isinstance(valor, (int, float)) else None
                por_pasada = (ram + ETIQUETAS_EXTRA) if ram is not None else 0

                a = avance.get(codigo, {})
                h1, h2 = a.get(1, 0), a.get(2, 0)
                estado = estado_de(h1, h2, por_pasada)
                faltan = max(por_pasada * PASADAS - (h1 + h2), 0)
                modulos_excel[codigo] = True

                hoja_pri.cell(row=fila, column=inicio, value=h1)
                hoja_pri.cell(row=fila, column=inicio + 1, value=h2)
                hoja_pri.cell(row=fila, column=inicio + 2, value=faltan)
                c = hoja_pri.cell(row=fila, column=inicio + 3, value=estado)
                c.fill = RELLENO_ESTADO[estado]
                hoja_pri.cell(row=fila, column=inicio + 4,
                              value=(a.get("ultima") or "")[:19].replace("T", " "))
            ancho(hoja_pri)
            print(f"Hoja '{hoja_pri.title}': {len(nuevas)} columnas anadidas")

    # ------------------------------------------------------- hoja de avance
    if "Avance rotulado" in wb.sheetnames:
        del wb["Avance rotulado"]
    ws = wb.create_sheet("Avance rotulado", 0)

    cols = ["ESTADO", "CODIGO_MODULO", "REGION", "FINCA", "RESPONSABLE", "RAMALES",
            "ETIQUETAS_POR_PASADA", "P1_GRABADAS", "P2_GRABADAS", "TOTAL_GRABADAS",
            "FALTAN", "AVANCE", "TELEFONOS", "REPETIDAS", "ULTIMA_FECHA"]
    ws.append(cols)
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center", wrap_text=True)

    # Todos los modulos del maestro, mas cualquiera con avance que ya no este en
    # el (un modulo retirado con etiquetas grabadas no puede desaparecer del
    # reporte: alguien fue y las pego).
    codigos = sorted(set(maestro) | set(avance))
    filas = []
    for codigo in codigos:
        m = maestro.get(codigo, {})
        ram = m.get("ramales")
        por_pasada = (ram + ETIQUETAS_EXTRA) if ram is not None else 0
        a = avance.get(codigo, {})
        h1, h2 = a.get(1, 0), a.get(2, 0)
        total = h1 + h2
        objetivo = por_pasada * PASADAS
        estado = estado_de(h1, h2, por_pasada)
        filas.append([
            estado, codigo, m.get("region", ""), m.get("finca", ""),
            m.get("responsable", ""), ram if ram is not None else "",
            por_pasada or "", h1, h2, total,
            max(objetivo - total, 0),
            round(total / objetivo, 4) if objetivo else "",
            len(a.get("telefonos", ())), repetidas.get(codigo, 0),
            (a.get("ultima") or "")[:19].replace("T", " "),
        ])

    # Primero lo que esta a medias, que es donde hay que volver; despues lo que
    # no se ha empezado; al final lo terminado.
    orden = {EN_PROCESO: 0, SIN_EMPEZAR: 1, COMPLETO: 2}
    filas.sort(key=lambda f: (orden[f[0]], -f[10], f[1]))

    i_avance = cols.index("AVANCE")
    for fila in filas:
        ws.append(fila)
        f = ws.max_row
        ws.cell(row=f, column=1).fill = RELLENO_ESTADO[fila[0]]
        ws.cell(row=f, column=i_avance + 1).number_format = "0%"
        if fila[cols.index("REPETIDAS")]:
            ws.cell(row=f, column=cols.index("REPETIDAS") + 1).fill = ROJO

    completos = sum(1 for f in filas if f[0] == COMPLETO)
    proceso = sum(1 for f in filas if f[0] == EN_PROCESO)
    sin_empezar = sum(1 for f in filas if f[0] == SIN_EMPEZAR)
    grabadas = sum(f[cols.index("TOTAL_GRABADAS")] for f in filas)
    objetivo = sum(f[cols.index("FALTAN")] for f in filas) + grabadas

    ws.append([])
    ws.append(["TOTAL", f"{len(filas)} modulos", "", "", "", "", "",
               sum(f[7] for f in filas), sum(f[8] for f in filas), grabadas,
               objetivo - grabadas,
               round(grabadas / objetivo, 4) if objetivo else "",
               "", sum(f[cols.index("REPETIDAS")] for f in filas), ""])
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)
    ws.cell(row=ws.max_row, column=i_avance + 1).number_format = "0%"

    ws.append([])
    ws.append([f"Completos: {completos}   En proceso: {proceso}   "
               f"Sin empezar: {sin_empezar}. "
               f"Un modulo esta completo solo con las DOS pasadas llenas."])
    ws.append([f"Avance leido del registro compartido el "
               f"{date.today():%Y-%m-%d}. Ramales tomados de modulos.json."])

    ws.freeze_panes = "A2"
    ancho(ws)

    try:
        wb.save(destino)
    except PermissionError:
        sys.exit(f"{destino.name} esta abierto en Excel. Cierralo y vuelve a correr esto.")

    print(f"\nCompletos    : {completos}")
    print(f"En proceso   : {proceso}")
    print(f"Sin empezar  : {sin_empezar}")
    print(f"Etiquetas    : {grabadas} de {objetivo} "
          f"({grabadas / objetivo:.1%})" if objetivo else "")
    print(f"\nGenerado: {destino}")


if __name__ == "__main__":
    main()
